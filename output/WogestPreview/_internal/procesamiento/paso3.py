"""
paso3.py - Módulo de Cruce de Datos
WOGest - Sistema de Validación de Renovaciones

Este módulo implementa la lógica de cruce de datos entre:
- Datos del Paso 1 (renovaciones validadas)
- Datos del Paso 2 (archivos WOQ procesados)

Funcionalidades:
- Correlación automática por número de WO
- Identificación de registros pendientes de cierre
- Generación de estadísticas de cruce
- Exportación de datos aptos para RPA
"""

import pandas as pd
import openpyxl
import os
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def realizar_cruce_datos(datos_paso1: List[Dict], datos_paso2: List[Dict]) -> Dict[str, Any]:
    """
    Realiza el cruce de datos entre Paso 1 (renovaciones) y Paso 2 (WOQ)
    
    Args:
        datos_paso1: Lista de registros validados del Paso 1
        datos_paso2: Lista de registros procesados del Paso 2
        
    Returns:
        Diccionario con resultados del cruce y estadísticas
    """
    try:
        logger.info("🔍 Iniciando cruce de datos...")
        
        if not datos_paso1 or not datos_paso2:
            return {
                'success': False,
                'message': 'No hay datos suficientes para realizar el cruce'
            }
        
        # Filtrar solo registros correctos del Paso 1
        registros_paso1_correctos = [
            r for r in datos_paso1 
            if r.get('estado', '').lower() == 'correcto'
        ]
        
        if not registros_paso1_correctos:
            return {
                'success': False,
                'message': 'No hay registros correctos en el Paso 1 para cruzar'
            }
        
        # Crear índice hash para búsqueda eficiente en datos del Paso 2
        # Normalizar claves para búsqueda (convertir a string y limpiar)
        indice_woq = {}
        for registro in datos_paso2:
            # Intentar diferentes campos que podrían contener el número de WO
            wo_numero = None
            for campo in ['N_WO', 'n_wo', 'N°_WO', 'wo', 'WO']:
                if campo in registro and registro[campo]:
                    wo_numero = str(registro[campo]).strip()
                    break
            
            if wo_numero:
                indice_woq[wo_numero] = registro
        
        logger.info(f"📊 Índice WOQ creado con {len(indice_woq)} registros")
        
        datos_cruzados = []
        estadisticas = {
            'total_paso1': len(registros_paso1_correctos),
            'total_cruzados': 0,
            'pendientes_cierre': 0,
            'cerrados': 0,
            'sin_woq': 0,
            'aptos_rpa': 0
        }
        
        # Procesar cada registro del Paso 1
        for registro_wo in registros_paso1_correctos:
            # Obtener número de WO del registro del Paso 1
            wo_numero = None
            for campo in ['wo', 'WO', 'n_wo', 'N_WO']:
                if campo in registro_wo and registro_wo[campo]:
                    wo_numero = str(registro_wo[campo]).strip()
                    break
            
            if not wo_numero:
                logger.warning(f"⚠️ Registro sin número de WO: {registro_wo}")
                continue
            
            # Buscar correspondencia en WOQ
            registro_woq = indice_woq.get(wo_numero)
            
            if registro_woq:
                # Registro encontrado en WOQ
                es_cerrado = registro_woq.get('ES_CERRADO', False)
                estado_cruce = 'Cerrado' if es_cerrado else 'Pendiente'
                apto_rpa = not es_cerrado  # Solo pendientes son aptos para RPA
                
                registro_cruzado = {
                    # Datos del WorkOrder (Paso 1)
                    **registro_wo,
                    
                    # Datos del WOQ (Paso 2) con prefijo para evitar conflictos
                    'woq_contrato': registro_woq.get('CONTRATO', ''),
                    'woq_n_wo': registro_woq.get('N_WO', ''),
                    'woq_cliente': registro_woq.get('CLIENTE', ''),
                    'woq_dealer': registro_woq.get('DEALER', ''),
                    'woq_status1': registro_woq.get('STATUS1', ''),
                    'woq_status2': registro_woq.get('STATUS2', ''),
                    'woq_cerrado': registro_woq.get('CERRADO', ''),
                    'woq_es_cerrado': es_cerrado,
                    'woq_f_sist': registro_woq.get('F_SIST', ''),
                    'woq_orden_contrato': registro_woq.get('ORDEN_CONTRATO', ''),
                    
                    # Campos de cruce
                    'estado_cruce': estado_cruce,
                    'apto_rpa': apto_rpa,
                    'timestamp_cruce': datetime.now().isoformat(),
                    'confianza_correlacion': 1.0  # Correlación exacta por WO
                }
                
                datos_cruzados.append(registro_cruzado)
                estadisticas['total_cruzados'] += 1
                
                if es_cerrado:
                    estadisticas['cerrados'] += 1
                else:
                    estadisticas['pendientes_cierre'] += 1
                    estadisticas['aptos_rpa'] += 1
                    
            else:
                # Registro sin WOQ correspondiente
                registro_cruzado = {
                    **registro_wo,
                    'estado_cruce': 'Sin WOQ',
                    'apto_rpa': False,
                    'timestamp_cruce': datetime.now().isoformat(),
                    'confianza_correlacion': 0.0
                }
                datos_cruzados.append(registro_cruzado)
                estadisticas['sin_woq'] += 1
        
        # Calcular porcentajes
        if estadisticas['total_paso1'] > 0:
            estadisticas['porcentaje_cruce'] = round(
                (estadisticas['total_cruzados'] / estadisticas['total_paso1']) * 100, 2
            )
            estadisticas['porcentaje_aptos_rpa'] = round(
                (estadisticas['aptos_rpa'] / estadisticas['total_paso1']) * 100, 2
            )
        else:
            estadisticas['porcentaje_cruce'] = 0
            estadisticas['porcentaje_aptos_rpa'] = 0
        
        mensaje_final = (
            f"Cruce completado exitosamente. "
            f"Procesados: {estadisticas['total_paso1']} registros, "
            f"Cruzados: {estadisticas['total_cruzados']}, "
            f"Aptos RPA: {estadisticas['aptos_rpa']}"
        )
        
        logger.info(mensaje_final)
        
        return {
            'success': True,
            'datos_cruzados': datos_cruzados,
            'estadisticas': estadisticas,
            'message': mensaje_final
        }
        
    except Exception as e:
        error_msg = f"Error en el cruce de datos: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {
            'success': False,
            'message': error_msg
        }

def exportar_datos_rpa(datos_cruzados: List[Dict], carpeta_destino: str) -> Dict[str, Any]:
    """
    Exporta datos aptos para RPA en formato Excel optimizado
    
    Args:
        datos_cruzados: Lista de registros cruzados
        carpeta_destino: Ruta de la carpeta donde guardar el archivo
        
    Returns:
        Diccionario con resultado de la exportación
    """
    try:
        logger.info(f"📊 Iniciando exportación RPA a: {carpeta_destino}")
        
        # Filtrar solo registros aptos para RPA
        registros_rpa = [
            r for r in datos_cruzados
            if r.get('apto_rpa', False)
        ]
        
        if not registros_rpa:
            return {
                'success': False,
                'message': 'No hay registros aptos para RPA'
            }
        
        # Crear carpeta si no existe
        os.makedirs(carpeta_destino, exist_ok=True)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"WOGest_RPA_Export_{timestamp}.xlsx"
        ruta_archivo = os.path.join(carpeta_destino, nombre_archivo)
        
        # Crear workbook de Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Datos_RPA"
        
        # Definir headers optimizados para RPA
        headers_rpa = [
            'WO', 'MANT', 'CLIENTE', 'REFERENCIA', 'TIPO', 'CANTIDAD',
            'PRECIO', 'CUOTA', 'TECNICO', 'PAGO',
            'WOQ_CONTRATO', 'WOQ_N_WO', 'WOQ_CLIENTE', 'WOQ_DEALER',
            'WOQ_STATUS1', 'WOQ_STATUS2', 'WOQ_CERRADO',
            'ESTADO_CRUCE', 'APTO_RPA', 'TIMESTAMP_PROCESAMIENTO',
            'CONFIANZA_CORRELACION'
        ]
        
        # Escribir headers con formato
        for col, header in enumerate(headers_rpa, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        # Escribir datos
        for row_idx, registro in enumerate(registros_rpa, 2):
            worksheet.cell(row=row_idx, column=1, value=registro.get('wo', ''))
            worksheet.cell(row=row_idx, column=2, value=registro.get('mant', ''))
            worksheet.cell(row=row_idx, column=3, value=registro.get('cliente', ''))
            worksheet.cell(row=row_idx, column=4, value=registro.get('referencia', ''))
            worksheet.cell(row=row_idx, column=5, value=registro.get('tipo', ''))
            worksheet.cell(row=row_idx, column=6, value=registro.get('cantidad', ''))
            worksheet.cell(row=row_idx, column=7, value=registro.get('precio', ''))
            worksheet.cell(row=row_idx, column=8, value=registro.get('cuota', ''))
            worksheet.cell(row=row_idx, column=9, value=registro.get('tecnico', ''))
            worksheet.cell(row=row_idx, column=10, value=registro.get('pago', ''))
            worksheet.cell(row=row_idx, column=11, value=registro.get('woq_contrato', ''))
            worksheet.cell(row=row_idx, column=12, value=registro.get('woq_n_wo', ''))
            worksheet.cell(row=row_idx, column=13, value=registro.get('woq_cliente', ''))
            worksheet.cell(row=row_idx, column=14, value=registro.get('woq_dealer', ''))
            worksheet.cell(row=row_idx, column=15, value=registro.get('woq_status1', ''))
            worksheet.cell(row=row_idx, column=16, value=registro.get('woq_status2', ''))
            worksheet.cell(row=row_idx, column=17, value=registro.get('woq_cerrado', ''))
            worksheet.cell(row=row_idx, column=18, value=registro.get('estado_cruce', ''))
            worksheet.cell(row=row_idx, column=19, value='SÍ' if registro.get('apto_rpa') else 'NO')
            worksheet.cell(row=row_idx, column=20, value=datetime.now().isoformat())
            worksheet.cell(row=row_idx, column=21, value=registro.get('confianza_correlacion', 0.0))
        
        # Aplicar formato a la tabla
        aplicar_formato_rpa(worksheet, len(registros_rpa))
        
        # Guardar archivo
        workbook.save(ruta_archivo)
        
        logger.info(f"✅ Exportación RPA completada: {len(registros_rpa)} registros")
        
        return {
            'success': True,
            'archivo': ruta_archivo,
            'total_registros': len(registros_rpa),
            'message': f'Exportación RPA completada: {len(registros_rpa)} registros exportados a {nombre_archivo}'
        }
        
    except Exception as e:
        error_msg = f"Error en exportación RPA: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {
            'success': False,
            'message': error_msg
        }

def aplicar_formato_rpa(worksheet, num_registros: int):
    """
    Aplica formato profesional a la hoja de Excel para RPA
    
    Args:
        worksheet: Hoja de trabajo de openpyxl
        num_registros: Número de registros de datos
    """
    try:
        # Ajustar ancho de columnas
        column_widths = {
            'A': 12,  # WO
            'B': 15,  # MANT
            'C': 25,  # CLIENTE
            'D': 20,  # REFERENCIA
            'E': 8,   # TIPO
            'F': 10,  # CANTIDAD
            'G': 12,  # PRECIO
            'H': 10,  # CUOTA
            'I': 15,  # TECNICO
            'J': 10,  # PAGO
            'K': 15,  # WOQ_CONTRATO
            'L': 12,  # WOQ_N_WO
            'M': 25,  # WOQ_CLIENTE
            'N': 15,  # WOQ_DEALER
            'O': 12,  # WOQ_STATUS1
            'P': 12,  # WOQ_STATUS2
            'Q': 10,  # WOQ_CERRADO
            'R': 15,  # ESTADO_CRUCE
            'S': 10,  # APTO_RPA
            'T': 20,  # TIMESTAMP
            'U': 15   # CONFIANZA
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Aplicar bordes a toda la tabla
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for row in range(1, num_registros + 2):  # +2 para incluir header y todos los datos
            for col in range(1, len(column_widths) + 1):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Aplicar formato alternado a las filas de datos
        light_fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row in range(2, num_registros + 2):  # Empezar desde fila 2 (después del header)
            if row % 2 == 0:  # Filas pares
                for col in range(1, len(column_widths) + 1):
                    worksheet.cell(row=row, column=col).fill = light_fill
        
        # Congelar primera fila (headers)
        worksheet.freeze_panes = 'A2'
        
        logger.info("✅ Formato aplicado a la hoja de Excel")
        
    except Exception as e:
        logger.warning(f"⚠️ Error al aplicar formato: {str(e)}")

def validar_integridad_cruce(registro_wo: Dict, registro_woq: Dict) -> Dict[str, Any]:
    """
    Valida la integridad de un cruce específico entre registros WO y WOQ
    
    Args:
        registro_wo: Registro del Paso 1 (WorkOrder)
        registro_woq: Registro del Paso 2 (WOQ)
        
    Returns:
        Diccionario con resultado de validación
    """
    validaciones = {
        'numero_orden_coincide': True,
        'cliente_consistente': True,
        'datos_completos': True
    }
    
    errores = []
    
    # Validar coincidencia de número de orden
    wo_numero = str(registro_wo.get('wo', '')).strip()
    woq_numero = str(registro_woq.get('N_WO', '')).strip()
    
    if wo_numero != woq_numero:
        validaciones['numero_orden_coincide'] = False
        errores.append(f"Números de orden no coinciden: {wo_numero} vs {woq_numero}")
    
    # Validar consistencia de cliente (si ambos tienen el campo)
    wo_cliente = str(registro_wo.get('cliente', '')).strip().upper()
    woq_cliente = str(registro_woq.get('CLIENTE', '')).strip().upper()
    
    if wo_cliente and woq_cliente and wo_cliente != woq_cliente:
        validaciones['cliente_consistente'] = False
        errores.append(f"Clientes no coinciden: {wo_cliente} vs {woq_cliente}")
    
    # Validar completitud de datos críticos
    campos_criticos_wo = ['wo', 'cliente', 'referencia', 'tipo']
    campos_criticos_woq = ['N_WO', 'CONTRATO', 'CLIENTE']
    
    for campo in campos_criticos_wo:
        if not registro_wo.get(campo):
            validaciones['datos_completos'] = False
            errores.append(f"Campo crítico faltante en WO: {campo}")
    
    for campo in campos_criticos_woq:
        if not registro_woq.get(campo):
            validaciones['datos_completos'] = False
            errores.append(f"Campo crítico faltante en WOQ: {campo}")
    
    # Calcular puntuación de confianza
    puntuacion_confianza = sum(validaciones.values()) / len(validaciones)
    
    return {
        'valido': all(validaciones.values()),
        'validaciones': validaciones,
        'errores': errores,
        'puntuacion_confianza': puntuacion_confianza
    }

def generar_reporte_cruce(datos_cruzados: List[Dict], estadisticas: Dict) -> str:
    """
    Genera un reporte textual del proceso de cruce
    
    Args:
        datos_cruzados: Lista de registros cruzados
        estadisticas: Estadísticas del cruce
        
    Returns:
        String con el reporte formateado
    """
    reporte = []
    reporte.append("=" * 60)
    reporte.append("REPORTE DE CRUCE DE DATOS - WOGEST")
    reporte.append("=" * 60)
    reporte.append(f"Fecha de procesamiento: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    reporte.append("")
    
    reporte.append("ESTADÍSTICAS GENERALES:")
    reporte.append(f"  • Total registros Paso 1: {estadisticas.get('total_paso1', 0)}")
    reporte.append(f"  • Registros cruzados exitosamente: {estadisticas.get('total_cruzados', 0)}")
    reporte.append(f"  • Registros sin WOQ: {estadisticas.get('sin_woq', 0)}")
    reporte.append(f"  • Porcentaje de cruce: {estadisticas.get('porcentaje_cruce', 0)}%")
    reporte.append("")
    
    reporte.append("ANÁLISIS DE ESTADOS:")
    reporte.append(f"  • Registros cerrados: {estadisticas.get('cerrados', 0)}")
    reporte.append(f"  • Registros pendientes: {estadisticas.get('pendientes_cierre', 0)}")
    reporte.append(f"  • Aptos para RPA: {estadisticas.get('aptos_rpa', 0)}")
    reporte.append(f"  • Porcentaje aptos RPA: {estadisticas.get('porcentaje_aptos_rpa', 0)}%")
    reporte.append("")
    
    # Análisis por cliente (top 10)
    if datos_cruzados:
        clientes = {}
        for registro in datos_cruzados:
            cliente = registro.get('cliente', 'Sin cliente')
            if cliente not in clientes:
                clientes[cliente] = {'total': 0, 'aptos_rpa': 0}
            clientes[cliente]['total'] += 1
            if registro.get('apto_rpa', False):
                clientes[cliente]['aptos_rpa'] += 1
        
        top_clientes = sorted(clientes.items(), key=lambda x: x[1]['total'], reverse=True)[:10]
        
        reporte.append("TOP 10 CLIENTES POR VOLUMEN:")
        for cliente, datos in top_clientes:
            reporte.append(f"  • {cliente}: {datos['total']} registros ({datos['aptos_rpa']} aptos RPA)")
    
    reporte.append("")
    reporte.append("=" * 60)
    
    return "\n".join(reporte)

# Funciones de utilidad adicionales

def buscar_registros_similares(wo_numero: str, datos_woq: List[Dict], umbral_similitud: float = 0.8) -> List[Dict]:
    """
    Busca registros similares cuando no hay coincidencia exacta
    Útil para identificar posibles errores de tipeo en números de WO
    """
    import difflib
    
    candidatos = []
    
    for registro in datos_woq:
        woq_numero = str(registro.get('N_WO', '')).strip()
        if woq_numero:
            similitud = difflib.SequenceMatcher(None, wo_numero, woq_numero).ratio()
            if similitud >= umbral_similitud:
                candidatos.append({
                    'registro': registro,
                    'similitud': similitud,
                    'wo_original': wo_numero,
                    'wo_candidato': woq_numero
                })
    
    return sorted(candidatos, key=lambda x: x['similitud'], reverse=True)

# Ejemplo de uso y pruebas
if __name__ == "__main__":
    # Datos de prueba
    datos_paso1_prueba = [
        {
            'wo': 'WO001',
            'cliente': 'Cliente A',
            'referencia': 'REF001',
            'tipo': 'AMCE',
            'cantidad': 1,
            'estado': 'correcto'
        },
        {
            'wo': 'WO002',
            'cliente': 'Cliente B',
            'referencia': 'REF002',
            'tipo': 'DMCE',
            'cantidad': -1,
            'estado': 'correcto'
        }
    ]
    
    datos_paso2_prueba = [
        {
            'N_WO': 'WO001',
            'CONTRATO': 'CONT001',
            'CLIENTE': 'Cliente A',
            'ES_CERRADO': False
        },
        {
            'N_WO': 'WO002',
            'CONTRATO': 'CONT002',
            'CLIENTE': 'Cliente B',
            'ES_CERRADO': True
        }
    ]
    
    # Ejecutar cruce de prueba
    resultado = realizar_cruce_datos(datos_paso1_prueba, datos_paso2_prueba)
    
    if resultado['success']:
        print("✅ Cruce de prueba exitoso")
        print(f"Registros cruzados: {len(resultado['datos_cruzados'])}")
        print(f"Estadísticas: {resultado['estadisticas']}")
        
        # Generar reporte
        reporte = generar_reporte_cruce(resultado['datos_cruzados'], resultado['estadisticas'])
        print("\n" + reporte)
    else:
        print(f"❌ Error en cruce de prueba: {resultado['message']}")

