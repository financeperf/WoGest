# paso4.py - Backend del Paso 4 (Exportación RPA)
# Requiere: paso3.realizar_cruce_datos y procesamiento.db_sqlite.*

from __future__ import annotations
import os, sys, logging
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

from paso3 import realizar_cruce_datos
from procesamiento.db_sqlite import leer_temp_paso1, leer_temp_paso2, limpiar_tablas_temporales

try:
    import webview
except Exception:
    webview = None

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger("paso4")

# ------------------------- helpers de normalización -------------------------

def _get_wo(row: Dict[str, Any]) -> str:
    for k in ("WO", "wo", "N_WO", "N°_WO", "N_WO2"):
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    return ""

def _get_orden_contrato(row: Dict[str, Any]) -> str:
    for k in ("ORDEN_CONTRATO", "CONTRATO", "woq_contrato", "WOQ_CONTRATO"):
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    return ""

def _val_es_cerrado(row: Dict[str, Any]) -> int:
    # normaliza a 0/1 leyendo variantes
    for k in ("es_cerrado", "ES_CERRADO", "woq_es_cerrado", "WOQ_ES_CERRADO", "woq_cerrado", "WOQ_CERRADO"):
        if k in row and row[k] is not None:
            v = str(row[k]).strip().upper()
            if v in ("1","SI","SÍ","TRUE","YES"): return 1
            if v in ("0","NO","FALSE"): return 0
            if v.isdigit(): return 1 if int(v)!=0 else 0
    return 0

def _val_estado_paso1(row: Dict[str, Any]) -> str:
    for k in ("Estado_Paso1","estado_paso1","ESTADO_PASO1"):
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    return ""

def _val_apto_rpa(row: Dict[str, Any]) -> str:
    for k in ("Apto RPA","APTO_RPA","apto_rpa"):
        if k in row and row[k] is not None:
            v = str(row[k]).strip().upper()
            if v in ("SI","SÍ","TRUE","YES","1"): return "SÍ"
            if v in ("NO","FALSE","0"): return "NO"
    return ""

# ------------------------- lectura/cruce -------------------------

def _cargar_cruce() -> Dict[str, Any]:
    df1 = leer_temp_paso1()
    df2 = leer_temp_paso2()
    datos_p1 = df1.to_dict(orient="records")
    datos_p2 = df2.to_dict(orient="records")
    res = realizar_cruce_datos(datos_p1, datos_p2)
    if not res.get("success"):
        return {"success": False, "message": res.get("message","Error en cruce")}
    return res

def _filtrar_para_exportar(datos_cruzados: List[Dict[str,Any]]) -> List[Dict[str,str]]:
    """
    Aplica tus condiciones y devuelve filas con SOLO columnas:
    - WO
    - ORDEN_CONTRATO
    Sin deduplicar.
    """
    out: List[Dict[str,str]] = []
    for r in datos_cruzados:
        if _val_es_cerrado(r) != 0:
            continue
        if _val_estado_paso1(r).upper() != "CORRECTO":
            continue
        if _val_apto_rpa(r) != "SÍ":
            continue
        wo = _get_wo(r)
        contrato = _get_orden_contrato(r)
        if wo or contrato:
            out.append({"WO": wo, "ORDEN_CONTRATO": contrato})
    return out

# ------------------------- exportación -------------------------

def _exportar_excel_wo_contrato(filas: List[Dict[str,str]], carpeta_destino: str) -> Dict[str,Any]:
    if not filas:
        return {"success": False, "message": "No hay registros que cumplan las condiciones."}

    os.makedirs(carpeta_destino, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre = f"RPA_WO_ORDEN_CONTRATO_{ts}.xlsx"
    ruta = os.path.join(carpeta_destino, nombre)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RPA"

    # headers
    ws.cell(1,1,"WO")
    ws.cell(1,2,"ORDEN_CONTRATO")
    for c in ("A1","B1"):
        ws[c].font = Font(bold=True, color="FFFFFF")
        ws[c].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # datos
    for i, fila in enumerate(filas, start=2):
        ws.cell(i,1, fila.get("WO",""))
        ws.cell(i,2, fila.get("ORDEN_CONTRATO",""))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28

    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for r in range(1, len(filas)+2):
        ws.cell(r,1).border = thin
        ws.cell(r,2).border = thin

    wb.save(ruta)

    return {
        "success": True,
        "archivo": ruta,
        "total_registros": len(filas),
        "message": f"Exportadas {len(filas)} filas (WO, ORDEN_CONTRATO) a {nombre}",
    }

# ------------------------- API para pywebview -------------------------

class Paso4API:
    """
    Métodos llamados desde export-utils.js y step4.js
    """

    def obtener_datos_para_rpa(self) -> Dict[str,Any]:
        try:
            cruce = _cargar_cruce()
            if not cruce.get("success"):
                return cruce
            datos = cruce.get("datos_cruzados", [])
            seleccion = _filtrar_para_exportar(datos)
            return {
                "success": True,
                "registros_rpa": datos,  # lo que pinta la tabla (si lo necesitas)
                "seleccion_exportable": seleccion,  # lo que realmente se exporta
                "estadisticas": {
                    "total_cruzados": len(datos),
                    "total_exportables": len(seleccion),
                },
                "total_registros": len(datos),
            }
        except Exception as e:
            log.exception("obtener_datos_para_rpa")
            return {"success": False, "message": str(e)}

    def exportar_excel_con_ruta(self, payload: Dict[str,Any]) -> Dict[str,Any]:
        """
        Punto de entrada usado por export-utils.js
        payload = {
            'datos': {...},            # ignorado/solo informativo
            'carpeta_destino': None|str,
            'nombre_archivo_original': str
        }
        """
        try:
            carpeta = payload.get("carpeta_destino") or os.path.abspath("exportables")

            cruce = _cargar_cruce()
            if not cruce.get("success"):
                return cruce

            datos = cruce.get("datos_cruzados", [])
            filas = _filtrar_para_exportar(datos)
            res = _exportar_excel_wo_contrato(filas, carpeta)
            if not res.get("success"):
                return res

            # limpieza completa tras exportar
            try:
                limpiar_tablas_temporales()
            except Exception as e:
                log.warning(f"No se pudieron limpiar temporales: {e}")

            # permitirá que el front decida volver al home
            res["redirect_home"] = True
            return res
        except Exception as e:
            log.exception("exportar_excel_con_ruta")
            return {"success": False, "message": str(e)}

    def seleccionar_directorio_exportacion(self) -> Dict[str,Any]:
        try:
            if webview and webview.windows:
                ruta = webview.windows[0].create_file_dialog(webview.FOLDER_DIALOG)
                if isinstance(ruta, list):
                    ruta = ruta[0] if ruta else None
            else:
                ruta = None
            if not ruta:
                return {"success": False, "message": "Selección cancelada"}
            return {"success": True, "ruta": ruta}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def abrir_carpeta_archivo(self, ruta_archivo: str) -> Dict[str,Any]:
        try:
            carpeta = os.path.dirname(ruta_archivo)
            if sys.platform.startswith("win"):
                os.startfile(carpeta)  # type: ignore
            elif sys.platform == "darwin":
                os.system(f'open "{carpeta}"')
            else:
                os.system(f'xdg-open "{carpeta}"')
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def limpiar_estado_completo(self) -> Dict[str,Any]:
        try:
            limpiar_tablas_temporales()
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

# Factory opcional
def create_api() -> Paso4API:
    return Paso4API()

if __name__ == "__main__":
    api = Paso4API()
    if webview:
        webview.create_window("Paso 4 (test)", html="<html><body>API Paso 4 lista</body></html>", js_api=api)
        webview.start()
    else:
        print(api.exportar_excel_con_ruta({"datos":{}, "carpeta_destino": "exportables"}))
